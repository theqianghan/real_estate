import requests
import webbrowser
from urllib.request import Request,urlopen
import urllib
import time
from bs4 import BeautifulSoup, SoupStrainer
import socks
import random
import pandas as pd
import os
import openpyxl
from itertools import cycle
import traceback
from lxml.html import fromstring

Malaysia_Housing_Data = os.getcwd() + "\\Malaysia_Housing_Data.xlsx"

if os.path.exists(Malaysia_Housing_Data):
	Data_File = openpyxl.load_workbook(Malaysia_Housing_Data)
else:
	wb = openpyxl.Workbook()
	dataset_worksheet_label = ['listing_id', 'listing_name',
                               'listing_address', 'township',
                               'state', 'property_type',
                               'rooms', 'bath', 'floor_area',
                               'floor_area_unit_measurement', 'land_size',
                               'quote_price']
	newsheet = wb.active
	newsheet.append(dataset_worksheet_label)
	wb.save(Malaysia_Housing_Data)
	Malaysia_Housing_Data = openpyxl.load_workbook(Malaysia_Housing_Data)


proxies = ['3.120.138.139:3128',
           '201.150.144.198:8080',
           '161.142.8.208',
           '45.160.34.97:8080',
           '202.162.222.154:51319',
           '134.209.30.49:3128',
           '223.204.218.239:8080',
           '134.209.24.90:3128']

#getting target listings from main sales site
link = "https://www.propertyguru.com.my/property-for-sale?status_code=ACT&market=residential&include_featured=1&_ref_section=ls&limit=30&_includePhotos=true&sort=date&order=desc"
headers = ['Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0',
           'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
           'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36',
           'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10; rv:33.0) Gecko/20100101 Firefox/33.0']


while True:
        try:
            initial_proxy = random.choice(proxies)
            print(initial_proxy)
            proxy_support = urllib.request.ProxyHandler({"http": initial_proxy,
                                                         "https": initial_proxy})
            opener = urllib.request.build_opener(proxy_support)
            urllib.request.install_opener(opener)
            req = urllib.request.Request(url=link, headers={'User-agent': random.choice(headers)})
            response = urllib.request.urlopen(req).read()
            if 'ROBOTS' not in str(results):
                break
            else:
                time.sleep(random.uniform(3,5))
                continue

        except:
            continue

#data wrangling to prepare for looped scraping
soup = BeautifulSoup(response, features = 'html.parser').find_all(href=True)
m = [i.attrs['href'] for i in soup]
res = set([y for y in m if 'property-listing' in y])
g = set([i.replace("#contact-agent","") for i in res])

#individual property listing parser
url_new = "http://www.propertyguru.com.my"

response_list = []


for i in g:
    while True:
        try:
            proxy_random = random.choice(proxies)
            proxy_random = random.choice(proxies)
            headers_random = random.choice(headers)
            print('HTTP Proxy: ' + proxy_random)
            print('HTTPS Proxy: ' + proxy_random)
            print('Faking the headers with ' + headers_random)
            print('Attempting...')
            proxy_support = urllib.request.ProxyHandler({"http": proxy_random,
                                                 "https": proxy_random})
            opener = urllib.request.build_opener(proxy_support)
            urllib.request.install_opener(opener)
            request_loop = urllib.request.Request(url=url_new + i, headers = {'User-agent': headers_random})
            results = urllib.request.urlopen(request_loop).read()
            if 'ROBOTS' not in str(results):
                print('MUAHAHA SUCCESS!!!')
                print('Appending now...')
                print('\n')
                response_list.append(results)
                time.sleep(random.uniform(8, 16))
                break
            else:
                print('ROBOT BLOCK FOUND. TRYING AGAIN...')
                print('\n')
                time.sleep(random.uniform(8,16))
                continue

        except:
            print('Some error came up with proxy. Trying again...')
            time.sleep(random.uniform(3, 5))
            print('\n')
            continue



soup_new = [BeautifulSoup(i, features = 'html.parser') for i in response_list]

#extractor
for i in range(0,len(soup_new)):
    listing_name = soup_new[i].find('div', class_="listing-title").text.replace(',','/')
    print('Listing Name' + ' : ' + listing_name)
    quote_price = str(soup_new[i].find_all('p')[7]).replace('<p>RM ','').replace(',','').replace('</p>','')
    print(float(quote_price))
    listing_id = soup_new[i].find('div', class_='listing-details-primary').find('div', class_='row').find_all('div', class_="col-xs-12 col-sm-6")[9].find('div',class_='value-block').text
    print(listing_id)
    listing_address = soup_new[i].find('span', itemprop = 'streetAddress').text.replace(',','/')
    print(listing_address)
    township = soup_new[i].find_all('li', itemtype = 'http://data-vocabulary.org/Breadcrumb')[3].find('span',itemprop='title').text
    print(township)
    city = soup_new[i].find_all('li', itemtype = 'http://data-vocabulary.org/Breadcrumb')[2].find('span',itemprop='title').text
    print(city)
    state = soup_new[i].find_all('li', itemtype = 'http://data-vocabulary.org/Breadcrumb')[1].find('span',itemprop='title').text
    print(state)
    description = soup_new[i].select_one('meta[name=description]')['content']
    print(description)
    property_type = soup_new[i].find('div', class_ = 'badge').text
    print(property_type)
    rooms = soup_new[i].find('span', itemprop = 'numberOfRooms').text
    print(int(rooms))
    bath = soup_new[i].find('div', class_ = 'property-info-element baths').find('span', class_ = "element-label").text
    print(int(bath))
    floor_area = soup_new[i].find('div', class_ = 'property-info-element area', itemprop = 'floorSize').select_one('meta[itemprop=value]')['content']
    print(int(floor_area))
    floor_area_unit_measurement = soup_new[i].find('div', class_ = 'property-info-element area', itemprop = 'floorSize').select_one('meta[itemprop=unitText]')['content']
    print(floor_area_unit_measurement)
    land_size = soup_new[i].find_all('div', class_ = 'col-xs-12 col-sm-6')[4].find('div', class_='property-attr', itemprop = 'additionalProperty').find('div', class_='value-block').text.replace(' sqft','')
    print(land_size)

    #here comes the feeding part
    item = [listing_id, listing_name, listing_address, township, state, property_type, rooms, bath, floor_area, floor_area_unit_measurement, land_size, quote_price]
    sheet = Data_File.active
    sheet.append(item)
    Data_File.save(os.getcwd() + "\\Malaysia_Housing_Data.xlsx")
